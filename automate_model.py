#!/usr/bin/env python
# coding: utf-8




import xlsxwriter
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Color, Fill, Font, Side, Border, Alignment, PatternFill
from openpyxl.cell import Cell
from openpyxl.utils import rows_from_range, cols_from_range, get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, utils
from openpyxl.utils import cols_from_range
from riskfree import rf_rate

def copy_range(range_str, src, dst):
    for row in rows_from_range(range_str):
        for cell in row:
            dst[cell].value = src[cell].value
    return

def format_selection(selection, sheet, style, font_a=Font(bold=True, size=12, color='FFFFFF')):
    for row in rows_from_range(selection):
        for cell in row:
            sheet[cell].style = style
            sheet[cell].font = font_a
            
def set_border(ws, cell_range, sty='thin'):
    rows = ws[cell_range]
    side = Side(border_style=sty, color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border
                
def fill_years_across(a,b,sheet):
    c=0
    for row in cols_from_range(f'{a}:{b}'):
            for cell in row:            
                sheet[cell] = year + c
                c += 1

# What is the company? Load Workbook
wb = openpyxl.load_workbook('BTS.xlsx')

company_country = 'thailand'

# How many years of projections?
nn = 8
projection = get_column_letter(7 + nn)
print(projection)

# When is the first year?
year = 2014

sheet = wb['Segments']

# Find Revenue Keyword to locate positions
df = pd.DataFrame()
for cellObj in list(sheet.columns)[0]:
    df = df.append(pd.Series(cellObj.value),ignore_index=True)
loco = df[df[0].str.contains('Revenues',na=False).values]
loco.index = loco.index + 1
beg = loco[0].index[0]
end = loco[0].index[1]


ws1 = wb.create_sheet('Revenue Model',0)
ws1.sheet_properties.tabColor = "1072BA"



# Copy first range of cells with copy_range function

first = 'A' + str(beg+1)
last = 'G' + str(end)
print(first, last)

copy_range(f'{first}:{last}',sheet,ws1)


ws1.move_range("A32:G34", rows=-20)

# Grab the segment numbers from the segment sheet

c = ws1[first].offset(column = 1).coordinate
d = ws1[first].offset(column = 6,row=(end-beg-1)).coordinate
print(c,d)

for row in cols_from_range(f'{c}:{d}'):
        for cell in row:  
            ws1[cell].value = f"={utils.quote_sheetname(sheet.title)}!{cell}"
            ws1[cell].number_format = '#,##0.0'

# Insert rows between segments, and add growth and percent of revenue

for i in range(end-beg):
    ws1.insert_rows(beg+2+3*i)
    ws1.insert_rows(beg+2+3*i)

for i in range(end-1-(beg)):
    ws1[f'A{beg+2+i*3}'] = '% Growth'
    ws1[f'A{beg+3+i*3}'] = '% of revenue'


ws1[ws1[f'{first}'].offset(row=-1).coordinate] = 'Sources of Revenue'
one = ws1[f'{first}'].offset(column = 1, row=-1).coordinate
two = ws1[f'{first}'].offset(column =6+nn,row=-1).coordinate
print(one)
print(two)

# Insert Years Across

for row in cols_from_range(f'{one}:{two}'):
        for cell in row:            
            ws1[cell] = year + c
            c += 1

# Moves the very last row (Revenue) to the top
seven = ws1[first].offset(row = ((end-beg -1) * 3)).coordinate
eight = ws1[first].offset(column = 6,row = ((end-beg -1) * 3)).coordinate

ws1.move_range(f"{seven}:{eight}", rows=-(int(seven[1:]) - 12))
print(seven,eight)


# Add the years across based off projection

ws1[ws1[f'{first}'].offset(row=-6).coordinate] = '(USD in Millions)'
three = ws1[f'{first}'].offset(column = 1, row=-6).coordinate
four = ws1[f'{first}'].offset(column = 6+nn,row=-6).coordinate
print(three)
print(four)

fill_years_across(three,four,ws1)

# Format the table 1 / 5 standard
past = 'Accent1'
future = 'Accent5'

a = ws1[f'{first}'].offset(row=-7).coordinate
b = ws1[f'{first}'].offset(column=6,row=-6).coordinate
print(a,b)

format_selection(f'{a}:{b}',ws1,past)

#TOTAL CHANGE#
c = ws1[f'{first}'].offset(row=((end-beg-1)*3)).coordinate
d = ws1[f'{first}'].offset(column=6,row=((end-beg-1)*3)).coordinate
print(c,d)
format_selection(f'{c}:{d}',ws1,past)

c = ws1[f'{first}'].offset(column=7,row=((end-beg-1)*3)).coordinate
d = ws1[f'{first}'].offset(column=6+nn,row=((end-beg-1)*3)).coordinate
print(c,d)
format_selection(f'{c}:{d}',ws1,future)


c = ws1[f'{first}'].offset(column=7,row=-1).coordinate
d = ws1[f'{first}'].offset(column=6+nn,row=-1).coordinate
print(c,d)
format_selection(f'{c}:{d}',ws1,future)

c = ws1[f'{first}'].offset(column=7,row=-7).coordinate
d = ws1[f'{first}'].offset(column=6+nn,row=-6).coordinate
print(c,d)
format_selection(f'{c}:{d}',ws1,future)

five = ws1[first].offset(row=-1).coordinate
six = ws1[first].offset(column = 6,row=-1).coordinate
print(five, six)
eighteen = (int(five[1:]) + 2)
nineteen = (int(five[1:]) + 3)

print(eighteen,nineteen)
format_selection(f'{five}:{six}',ws1,past)

for i in range(end-beg-1):
    format_selection(f'A{eighteen+i*3}:{projection}{nineteen+i*3}',ws1,'20 % - Accent1', None) 

seven = ws1[first].offset(column = 1,row = -4).coordinate
eight = ws1[first].offset(column = (6+nn),row = -4).coordinate
for row in cols_from_range(f'{seven}:{eight}'):
        for cell in row:  
            before = ws1[cell].offset(column = -1,row=-1).coordinate
            above = ws1[cell].offset(row=-1).coordinate
            ws1[cell] = f"=({above} - {before})/{before}"
            ws1[cell].number_format = '0.00%'
print(seven, eight)

# Calculate Growth Rate

for i in range(end-beg-1):
    for row in cols_from_range(f'B{18+i*3}:G{18+i*3}'):
        for cell in row:  
            before = ws1[cell].offset(column = -1,row=-1).coordinate
            above = ws1[cell].offset(row=-1).coordinate
            ws1[cell] = f"=({above}-{before})/{before}"
            ws1[cell].number_format = '0.00%'

# Calculate Percentage of Revenue

for i in range(end-beg-1):
    for row in cols_from_range(f'B{19+i*3}:{projection}{19+i*3}'):
        for cell in row:  
            total = ws1[cell].offset(row=-7).coordinate
            above = ws1[cell].offset(row=-2).coordinate
            col = ws1[cell].coordinate[0]
            
            ws1[cell] = f"=({above})/{str(col)+'12'}"
            ws1[cell].number_format = '0.00%'

# Project out revenues from previous and growth rate

for i in range(end-beg-1):
    for row in cols_from_range(f'H{17+i*3}:{projection}{17+i*3}'):
        for cell in row:  
            a = (ws1[cell].offset(column=-1).coordinate)
            b = (ws1[cell].offset(row=1).coordinate)
            ws1[cell] = f"={a}*(1+{b})"
            ws1[cell].number_format = '#,##0.0'

# Locate key points to create borders

a = ws1[first].offset(row=-1).coordinate
b = ws1[first].offset(row=(((end-beg-1)*3)-1),column=6).coordinate
c = ws1[first].offset(row=(((end-beg-1)*3)),column=6).coordinate
d = b[1:]
e = c[1:]
f = a[1:]
print(a,b,c,d,e,f)

# Create borders

set_border(ws1,'A10:G13')
set_border(ws1,f'{a}:{b}')
set_border(ws1,f'{a}:{c}')
set_border(ws1,f'H10:{projection}13')
set_border(ws1,f'H{f}:{projection}{d}')
set_border(ws1,f'H{e}:{projection}{e}')

# Create Total Revenue at bottom from adding up segments

ws1[first].offset(row=((end-beg -1) * 3)).value = 'Total Revenue'
ws1['A13'] = "% Growth"
# Reference the B and L
a = ws1[first].offset(row=((end-beg -1) * 3)).coordinate[1:]
b = (int(a)-3)
for row in cols_from_range(f'B{a}:{projection}{a}'):
        for cell in row:
            col = ws1[cell].coordinate[0]
            string = ''
            for i in range(end-beg-1):
                string += (str(col+str(b-i*3)+" + "))

            ws1[cell] = f"={string[:-3]}"
            ws1[cell].number_format = '#,##0.0'
print(string[:-3])
print(a,b)

# Add Key Cells, format
# Locate Name of Company and add to Title Sheet
name = sheet['A5'].value.split('>')[0][:-1]
ws1['A7'].value = name
ws1['A7'].font = Font(bold=True,sz=14)

nine = ws1[first].offset(column = 1,row = 1).coordinate
ws1['B13'] = ""
ws1['A13'] = "% Growth"
ws1['A8'] = 'Revenue Model'
ws1['A8'].font = Font(bold=True,sz=14)

for i in range(end-beg-1):
    for row in cols_from_range(f'B{18+i*3}:B{18+i*3}'):
        for cell in row:
            ws1[cell].value = ""

ws1['H10'].value = 'Projections'
ws1.merge_cells(f'H10:{projection}10')
ws1['H10'].alignment = Alignment(horizontal='center', vertical='center')

# Make Total Revenues match with the bottom
for row in cols_from_range(f'H12:{projection}12'):
        for cell in row:
            col = ws1[cell].coordinate[0]
#             print(col)
            ws1[cell].value = f"={col+str(17+3 * (end-beg-1))}"
            ws1[cell].number_format = '#,##0.0'

seven = ws1[first].offset(row = ((end-beg -1) * 3)).coordinate
nine = ws1[first].offset(row = ((end-beg-1)*3 + 3)).coordinate
ten = ws1[first].offset(row = ((end-beg-1)*3 + 3 +(end-beg-2))).coordinate
eleven = ws1[first].offset(row = ((end-beg-1)*3 + 2)).coordinate
twelve = ws1[first].offset(column=1, row = ((end-beg-1)*3 + 2)).coordinate
thirt = ws1[first].offset(column = 1, row = ((end-beg-1)*3 + 3)).coordinate
fourt = ws1[first].offset(column = 1, row = ((end-beg-1)*3 + 3 +(end-beg-2))).coordinate

# Grab Names of Segment from Above Table
a = 0
for row in cols_from_range(f'{nine}:{ten}'):
        for cell in row:  
            print(cell)
            ws1[cell].value = f'=A{str(17+a)}'
            a += 3
            
# Create Average Growth Table
ws1[eleven] = 'Average Growth'
ws1.merge_cells(f'{eleven}:{twelve}')
ws1[eleven].alignment = Alignment(horizontal='center')
format_selection(f'{eleven}:{twelve}',ws1, past)
a = 0
for row in cols_from_range(f'{thirt}:{fourt}'):
        for cell in row:
            ws1[cell].value = f'=AVERAGE(C{18+a}:G{18+a})'
            ws1[cell].number_format = '0.00%'
            a += 3
            
# Set border around Average Table         
set_border(ws1,f'A{(int(seven[1:])+2)}:B{str((int(seven[1:])+2) + (end-beg-1))}')
print(nine,ten,thirt,fourt)

######### DCF ###########################################
 
ws2 = wb.create_sheet('DCF',0)
ws2.sheet_properties.tabColor = "7CC025"

ws2['B3'] = name
ws2['B3'].font = Font(bold=True,sz=14)
ws2['B4'] = 'Discounted Cash Flow Model'
ws2['B4'].font = Font(bold=True,sz=12)

sheet1 = wb['Income Statement']
b = pd.DataFrame(sheet1.values)
b.index = b.index + 1
print(b)

# Find First Group of Rows (SGA, Other Op. Inc.)

list1 = ['  Gross Profit','  Gross Margin']
s2 = b.loc[b[0].isin(list1)].index[0] + 2
st1 = f'A{str(s2)}' 

print(st1)
print(s2)
print(sheet1['A25'].value)

na = b.loc[b[0].isna()]
e = na.loc[na.index > s2].index[0] - 1
en = f'G{str(e)}'
print(e,en)

def copy_from_formula(st,en,first, second, shift=0):
    for row in cols_from_range(f'{st}:{en}'):
            for cell in row:
                loca = cell[0] + str(int(cell[1:]) - shift)
                second[loca].value = f"={utils.quote_sheetname(first.title)}!{cell}"
                second[loca].number_format = '#,##0.0'

# Copy Formula SG&A to Other Operating Exp
                
copy_from_formula(st1,en,sheet1,ws2)

en2 = f'G{str(e+1)}'
en1 = f'A{str(e+1)}'
print(en1,en2)

# Set Grey Bar at Top
format_selection(f'A16:{projection}16',ws2, '20 % - Accent3', font_a=Font(bold=False, size=12))
  

# Format Blue Bar
format_selection(f'A17:G18',ws2, past, font_a=Font(bold=False, size=12, color='FFFFFF'))


# Format Light Blue Bar
format_selection(f'H17:{projection}18',ws2, future, font_a=Font(bold=False, size=12, color='FFFFFF'))



def find_block_loc(list_words,b):
    s = b.loc[b[0].isin(list_words)].index[0] + 2
    st = f'A{str(s)}' 
    
    na = b.loc[b[0].isna()]
    e = na.loc[na.index > s].index[0] - 1
    en = f'G{str(e-1)}'
    
    return (st, en)

# Income to Other 

st5, en5 = find_block_loc(['  Net Interest Exp.'],b)
print(st5,en5)
s = int(st5[1:]) - e - 2
print(f's: {s}')
copy_from_formula(st5,en5,sheet1,ws2,s)

# Restruct to Other Unusual

st3, en3 = find_block_loc(['  EBT Excl. Unusual Items'], b)
print(st3,en3)

start = (int(st3[1:]))
last = int(en5[1:]) - s + 1
s1 = start - last
print(start, last, s1)

copy_from_formula(st3,en3,sheet1,ws2,s1)

# Interest Exp/ Interest Inv. Income

st2, en2 = find_block_loc(['  Operating Income'], b)
print(st2,en2)
s = int(en3[1:]) - s1 - int(st2[1:]) + 2
print(f's: {s}')
copy_from_formula(st2,en2,sheet1,ws2,-s)

# Income Tax Line

st4, en4 = find_block_loc(['  EBT Incl. Unusual Items'], b)
print(st4,en4)

eend = int(en2[1:]) + s + 2
sstart = int(en4[1:]) 
s = sstart - eend
print(s)

copy_from_formula(st4,en4,sheet1,ws2,s)

# Insert Rows

for i in range((int(en4[1:])-s) - int(st1[1:])):
    ws2.insert_rows(int(st1[1:])+1+2*i)

# Find last line/row in table
df1 = pd.DataFrame()
for cellObj in list(ws2.columns)[0]:
    df1 = df1.append(pd.Series(cellObj.value),ignore_index=True)
# df1[df1[0].str.contains('Income Tax Expense',na=False).values]
loca2 = df1.index[-1] + 4
print(loca2)

# Repeat for 27 times from end - beg, Replace 70 with a reference
# Every other background color

for i in range(int((loca2-16)/2)):
    format_selection(f'A{18+ 2 * i}:G{18+ 2 * i}',ws2, '20 % - Accent3', font_a=Font(bold=False, size=12))

for i in range(int((loca2-16)/2)):
    format_selection(f'H{18+ 2 * i}:{projection}{18+ 2 * i}',ws2, '20 % - Accent5', font_a=Font(bold=False, size=12))

format_selection(f'A23:G24',ws2, past, font_a = Font(bold=False, size=12, color='FFFFFF'))
format_selection(f'H23:{projection}24',ws2, future,font_a=Font(bold=False, size=12, color='FFFFFF'))

# Place Operating Income in Cell A35 // Change the 33, needs to reference
# e - s2 + 1

begin =int(st1[1:])
#               SG&A to Other Op     Income to Other Unusual      #
move = (int(en[1:]) - int(st1[1:]) + 1) * 2
print(move)
ws2[f'A{begin + move}'].value = 'Operating Income'
ws2[f'A{begin + move + 1}'].value = '% of Revenue'
print((begin + move),(begin+move+1))

# Format Op Income

format_selection(f'A{begin+move}:G{begin+move+1}',ws2, past, font_a = Font(bold=False, size=12, color='FFFFFF'))
format_selection(f'H{begin+move}:{projection}{begin+move+1}',ws2, future, font_a=Font(bold=False, size=12, color='FFFFFF'))

# Calculate Operating Income Formula
start = 25 + (e - s2) * 2
print(f'start: {start}')
for row in cols_from_range(f'B{begin + move}:{projection}{begin+ move}'):
    for cell in row:
        col = ws2[cell].coordinate[0]
        string = ''
        for i in range(e-s2+1):
            string += (str(col+str(start-i*2)+" - "))
        ws2[cell].value = f"={col}23 - {string[:-3]}"
        ws2[cell].number_format = '#,##0.0'
print(string)
        
for row in cols_from_range(f'B{begin + move+1}:{projection}{begin+ move+1}'):
    for cell in row:
        col = ws2[cell].coordinate[0]
        op = ws2[cell].offset(row=-1).coordinate
        ws2[cell].value = f'={op}/{col}19'
        ws2[cell].number_format = '0.00%'
        
f = df1.index[-1] - 1
g = df1.index[-1]
print(f,g)

# format EBT
format_selection(f'A{f}:G{g}',ws2, past, font_a = Font(bold=False, size=12, color='FFFFFF'))
format_selection(f'H{f}:{projection}{g}',ws2, future, font_a=Font(bold=False, size=12, color='FFFFFF'))
ws2[f'A{f}'].value = 'Earnings Before Tax (EBT)'
ws2[f'A{g}'].value = '% of Revenue'

# Locate points // Formula error

first = int(en3[1:]) - int(st3[1:]) + 1
second = int(en5[1:]) - int(st5[1:]) + 1
third = (first + second + (int(en[1:]) - int(st1[1:]) + 1)) * 2
print(third)

# Format EBIT

format_selection(f'A{begin+third+2}:G{begin+third+3}',ws2, past, font_a = Font(bold=False, size=12, color='FFFFFF'))
format_selection(f'H{begin+third+2}:{projection}{begin+third+3}',ws2, future, font_a=Font(bold=False, size=12, color='FFFFFF'))
     
# Add Earnings Before Interest & Taxes (EBIT)

ws2[f'A{begin + third + 2}'].value = 'Earnings Before Interest & Taxes (EBIT)'
ws2[f'A{begin + third + 3}'].value = '% of Revenue'
print((begin+third+2),(begin+third+3))

# Input Formula for EBIT //  Error found in this line
for row in cols_from_range(f'B{begin + third + 2}:{projection}{begin+ third + 2}'):
    for cell in row:
        col = ws2[cell].coordinate[0]
        string = ''
        for i in range(s-2):
            string += (str(col+str((begin+third)-i*2)+" + "))
        ws2[cell].value = f"={col}{begin+move} + {string[:-3]}"
        ws2[cell].number_format = '#,##0.0'


        
for row in cols_from_range(f'B{begin + third + 3}:{projection}{begin+ third + 3}'):
    for cell in row:
        col = ws2[cell].coordinate[0]
        op = ws2[cell].offset(row=-1).coordinate
        ws2[cell].value = f'={op}/{col}19'
        ws2[cell].number_format = '0.00%'

# Put the title above income statement

inp = ['Gross Margin','Gross Profit','% of Revenue','Cost of Revenues','% YoY Growth','Total Revenue','USD','Discounted Cash Flow Model','Discount Period']
a=1
for i in inp:
    ws2['A' + str(int(st1[1:]) - a)].value = i
    a+=1

# Add Discount Period to the top of table    
b=1    
for row in cols_from_range(f'H16:{projection}16'):
    for cell in row:
        ws2[cell].value = b
        b+=1

# Calculate gross margin formula

loca = str(int(st1[1:]) - 1)
fmt = '0.00%'
print(loca)
def fill_from_loc_ref(n, eq, form,loc1,loc2):
    for row in cols_from_range(f'B{n}:{projection}{n}'):
        for cell in row:
            one = ws2[cell].offset(row=loc1).coordinate
            two = ws2[cell].offset(row=loc2).coordinate
            ws2[cell] = f"={one}{eq}{two}"
            ws2[cell].number_format = form
            
fill_from_loc_ref(loca,'/',fmt,-1,-5)

loca = str(int(st1[1:]) - 2)
print(loca)
fmt = '#,##0.0'
fill_from_loc_ref(loca,'+',fmt,-2,-4)

loca = str(int(st1[1:]) - 3)
print(loca)
fmt = '0.00%'
fill_from_loc_ref(loca,'/',fmt,-1,-3)

# Copy Cost of Revenues

for row in cols_from_range(f'B22:G22'):
            for cell in row:
                loca = cell[0] + str(int(cell[1:]) - 1)
                ws2[loca].value = f"=-{utils.quote_sheetname(sheet1.title)}!{cell}"
                ws2[loca].number_format = '#,##0.0'

def fill_from_loc_ref1(n, eq, form,loc1,loc2):
    for row in cols_from_range(f'B{n}:{projection}{n}'):
        for cell in row:
            one = ws2[cell].offset(row=loc1).coordinate
            two = ws2[cell].offset(row=loc2,column=-1).coordinate
            ws2[cell] = f"={one}{eq}{two}-1"
            ws2[cell].number_format = form
loca = str(int(st1[1:]) - 5)
fmt = '0.00%'
print(loca)
fill_from_loc_ref1(loca,'/',fmt,-1,-1)

# Forecast cost of revenues

for row in cols_from_range(f'H21:{projection}21'):
    for cell in row:  
        before = ws2[cell].offset(column=-1).coordinate
        below = ws2[cell].offset(row=1).coordinate
        ws2[cell] = f"={before}*(1+{below})"
        ws2[cell].number_format = '#,##0.0'

# Can be used for Moving Average

for row in cols_from_range(f'H22:{projection}22'):
    for cell in row:  
        before = ws2[cell].offset(column=-1).coordinate
        after = ws2[cell].offset(column=-4).coordinate
        ws2[cell] = f"=AVERAGE({before}:{after})"
        ws2[cell].number_format = '0.00%'

# change input to reference from location
copy_from_formula('B12',f'{projection}12',ws1,ws2,-7)

fill_years_across('B18',f'{projection}18',ws2)

# Add percent of revenue
for i in range(int(en[1:]) - int(st1[1:]) + 1):
    print(st1[0] + str(int(st1[1:])+1 +2 * i))
    ws2[(st1[0] + str(int(st1[1:])+1 +2 * i))].value = '% of Revenue'

# Calculate percentage of revenue across table

for i in range(int(en[1:]) - int(st1[1:]) + 1):
    for row in cols_from_range(f'B{26+i*2}:G{26+i*2}'):
        for cell in row:  
            total = ws2[cell].offset(row=-(7+i*2)).coordinate
            above = ws2[cell].offset(row=-1).coordinate
            col = ws2[cell].coordinate[0]
            
#             ws2[cell] = f"=({above})/{str(col)+'12'}"
            ws2[cell] = f"=IF(ISERROR({above}/{total}),0.0,{above}/{total})"
            ws2[cell].number_format = '0.00%'
        
# Calculate forecast top

for i in range(int(en[1:]) - int(st1[1:]) + 1):
    for row in cols_from_range(f'H{25+i*2}:{projection}{25+i*2}'):
        for cell in row:  
            before = ws2[cell].offset(column=-1).coordinate
            growth = ws2[cell].offset(row=1).coordinate
            col = ws2[cell].coordinate[0]

            ws2[cell] = f"=IF(ISERROR({before}*(1+{growth})),0.0,{before}*(1+{growth}))"
            ws2[cell].number_format = '#,##0.0'

# Input title, % of Revenue // delete print statement

for i in range(int(st3[1:]) - int(st2[1:]) + 4):
    print(st2[0] + str(int(st2[1:])+2 +2 * i))
    ws2[(st2[0] + str(int(st2[1:])+2 +2 * i))].value = '% of Revenue'

# Calculate % of Revenue Again (Bottom)
loc1 = (s2 +((e-s2+1)* 2 + 3))
print(loc1)
for i in range(int(st3[1:]) - int(st2[1:])+ 2):
    for row in cols_from_range(f'B{loc1+i*2}:G{loc1+i*2}'):
        for cell in row:  
            total = ws2[cell].offset(row=-(19+i*2)).coordinate
            above = ws2[cell].offset(row=-1).coordinate
            col = ws2[cell].coordinate[0]

            ws2[cell] = f"=IF(ISERROR(({above})/{total}),0.0,{above}/{total})"
            ws2[cell].number_format = '0.00%'

# Calculate Forecast (bottom)
loc2 = (s2 +((e-s2+1)* 2 + 2))
print(loc2)
for i in range(s- 1):
    for row in cols_from_range(f'H{loc2+i*2}:{projection}{loc2+i*2}'):
        for cell in row:  
            before = ws2[cell].offset(column=-1).coordinate
            growth = ws2[cell].offset(row=1).coordinate
            col = ws2[cell].coordinate[0]
            
            ws2[cell] = f"=IF(ISERROR({before}*(1+{growth})),0.0,{before}*(1+{growth}))"
            ws2[cell].number_format = '#,##0.0'

format_selection(f'A{loca2-1}:G{loca2}',ws2, past, font_a = Font(bold=False, size=12, color='FFFFFF'))
format_selection(f'H{loca2-1}:{projection}{loca2}',ws2, future, font_a=Font(bold=False, size=12, color='FFFFFF'))

ws2[f'A{loca2-1}'].value = 'Net Income'
ws2[f'A{loca2}'].value = 'Net Margin'

#Input growth formula for Interest Expense
for i in range(2):
    for row in cols_from_range(f'H{(f-4)+i*2}:{projection}{(f-4)+i*2}'):
            for cell in row:  
                before = ws2[cell].offset(column=-1).coordinate
                growth = ws2[cell].offset(row=1).coordinate
                col = ws2[cell].coordinate[0]

                ws2[cell] = f"=IF(ISERROR({before}*(1+{growth})),0.0,{before}*(1+{growth}))"
                ws2[cell].number_format = '#,##0.0'

print(f,g)
# Calculate EBT 
for row in cols_from_range(f'B{f}:{projection}{f}'):
            for cell in row:  
                col = ws2[cell].coordinate[0]
                string = ''
                for i in range(3):
                    string += (str(col+str((begin+third+2)+i*2)+" + "))

                ws2[cell] = f"={string[:-3]}"
                ws2[cell].number_format = '#,##0.0'
print(string[:-3])
# EBT % of Revenue
for i in range(2):
    for row in cols_from_range(f'B{g+i*2}:{projection}{g+i*2}'):
        for cell in row:  
            above = ws2[cell].offset(row=-1).coordinate
            col = ws2[cell].coordinate[0]

            ws2[cell] = f"=IF(ISERROR({above}/{col}19),0.0,{above}/{col}19)"
            ws2[cell].number_format = '0.00%'



# Net Income Calculation
for row in cols_from_range(f'B{loca2-1}:{projection}{loca2-1}'):
        for cell in row:  
            tax = ws2[cell].offset(row=-2).coordinate
            ebt = ws2[cell].offset(row=-4).coordinate
            col = ws2[cell].coordinate[0]

            ws2[cell] = f"=IF(ISERROR({ebt}-{tax}),0.0,{ebt}-{tax})"
            ws2[cell].number_format = '#,##0.0'
            
#Net Margin Calculation
for row in cols_from_range(f'B{loca2}:{projection}{loca2}'):
        for cell in row:  
            above = ws2[cell].offset(row=-1).coordinate
            col = ws2[cell].coordinate[0]

            ws2[cell] = f"=IF(ISERROR({above}/{col}19),0.0,{above}/{col}19)"
            ws2[cell].number_format = '0.00%'

set_border(ws2,f'A16:{projection}16')
set_border(ws2,f'A17:{projection}{loca2}')
set_border(ws2,f'A17:G{loca2}')

ws2['H17'].value = 'Projections'
ws2.merge_cells(f'H17:{projection}17')
ws2['H17'].alignment = Alignment(horizontal='center', vertical='center')


################## DCF Assumptions ###############################
##################################################################


ws3 = wb.create_sheet('DCF Assumptions',0)
ws3.sheet_properties.tabColor = "D2FF98"
sheet3 = wb['Balance Sheet']

ws3['C2'] = name
ws3['C2'].font = Font(bold=True,sz=14)

ws3['C4'].value = 'Discounted Free Cash Flow Assumptions'
ws3.merge_cells(f'C4:F4')
format_selection(f'C4:F4',ws3, past, font_a = Font(bold=False, size=11, color='FFFFFF'))
ws3['C4'].alignment = Alignment(horizontal='center', vertical='center')


format_selection(f'E14:F16',ws3, past, font_a = Font(bold=False, size=11, color='FFFFFF'))
ws3['E14'] = 'Implied Price'
ws3['E15'] = 'Current Price'
ws3['E16'] = 'Under/Over Valued'

format_selection(f'C18:E18',ws3, past, font_a = Font(bold=False, size=11, color='FFFFFF'))
ws3['C18'] = 'Final Valuation'
ws3['D18'] = 'Implied Price'
ws3['E18'] = 'Weighting'
ws3['C19'] = 'DCF'
ws3['C20'] = 'COMP'

ws3['C21'] = 'Implied Price'
ws3['C22'] = 'Current Price'
ws3['C23'] = 'Under/Over Valued'
format_selection(f'C21:E23',ws3, past, font_a = Font(bold=False, size=11, color='FFFFFF'))

assumptions = ['Tax Rate','Risk Free Rate','Beta','Total Equity Risk Premium','Cost of Equity (Re)','Cost of Debt (Rd)','% Debt',' % Equity','WACC']
assumptions2 = ['Terminal Growth Rate','Debt','Excess Cash', 'Net Debt','Market Capitalization','Debt + Equity','Calculated EV','Implied Equity','Shares Outstanding']
for i,j in enumerate(assumptions):
    ws3[f'C{5+i*1}'].value = j
        
for i,j in enumerate(assumptions2):
        ws3[f'E{5+i*1}'].value = j
        
set_border(ws3,'C4:F4','medium')
set_border(ws3,'C5:F13','medium')
set_border(ws3,'E14:F16','medium')
set_border(ws3,'C18:E23','medium')

ws3['D6'].value = rf_rate(company_country)
ws3['D6'].number_format = '0.00%'


# Shares Outstanding formula
for i,j in enumerate(sheet3["A"]):
    if j.value == 'Total Shares Out. on Filing Date':
        ws3['F13'].value = f"={utils.quote_sheetname(sheet3.title)}!G{i+1}"
ws3['F10'] = '=F8 + F9'
ws3['F9'] = '=F13 * D22'

# Read in CSV's and locate industry

firms = pd.read_csv('firms.csv')
firms.index = firms.index +2
betas = pd.read_csv('industry_beta.csv')
betas.index = betas.index +2
industry = firms[firms['Company Name']==name]['Industry Group'].values[0]
print(industry)

# Locate firm beta from industry 

firm_beta = betas[betas['Industry Name'] == industry]['Unlevered beta'].values[0]
ws3['D7'].value = firm_beta
ws3['D7'].number_format = '0.00'

print(firm_beta)


wb.save('copy.xlsx')





